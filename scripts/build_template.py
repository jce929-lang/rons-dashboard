"""Build the seeded Google Sheets template (Quotes, Viability, Sales, Config)."""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "Rons Dashboard Template.xlsx"
SRC = Path(__file__).resolve().parent.parent / "USL Sales Template and Forecast.xlsx"

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 40)


wb = Workbook()

# --- Quotes ---
ws = wb.active
ws.title = "Quotes"
ws.append(["quote", "author"])
quotes = [
    ("Grace is the engine for growth", ""),
    ("Be on time, try your best, act right", "Gene Keady"),
    ("Experience is a hard teacher — she gives the test first, then the lesson", "Vern Law"),
    ("Excellence honors God and inspires people. If you will do the best you can, so will they", ""),
    ("If you're not unique, you'd better be cheap", ""),
    ("Life is hard; it's harder if you're stupid", ""),
    ("Bad habits are made in good times, good habits are made in bad times", ""),
    ("It's wise to be fearful when others are greedy and greedy when others are fearful", "Warren Buffett"),
    ("…that if any would not work, neither should he eat", "II Thessalonians 3:10"),
    ("When the going gets tough, the tough get going", ""),
    ("Faith doesn't get you around trouble, it gets you through it", ""),
    ("You will never reach your destination if you stop and throw stones at every dog that barks", "Winston Churchill"),
]
for q in quotes:
    ws.append(q)
style_header(ws)
autosize(ws)

# --- Viability ---
ws = wb.create_sheet("Viability")
ws.append(["product", "potential", "channels", "remaining_work"])
viability = [
    ("Handgun Locker", 10, 10, 7),
    ("Jeep/Bronco Locker", 10, 10, 1),
    ("Stryker Case/Trays", 10, 4, 7),
    ("Stryker Latch", 10, 3, 6),
    ("Home Gun Locker", 8, 10, 8),
    ("Water Smoker", 8, 2, 5),
    ("Underseat Truck Locker Core", 9, 10, 10),
    ("Reclining Creeper", 9, 2, 6),
    ("Trimmer Guard", 6, 4, 2),
    ("Mobile Cubicles", 3, 3, 10),
    ("Knee Rehab", 8, 3, 7),
]
for row in viability:
    ws.append(row)
style_header(ws)
autosize(ws)

# --- Config ---
ws = wb.create_sheet("Config")
ws.append(["key", "value"])
config = [
    ("usl_unit_price", 795),
    ("fol_unit_price", 1500),
    ("ram_unit_price", 795),
    ("dashboard_title", "Ron's Dashboard"),
]
for row in config:
    ws.append(row)
style_header(ws)
autosize(ws)

# --- Sales --- seed from existing xlsx Data and Forecast tab
ws = wb.create_sheet("Sales")
columns = [
    "week_of", "type",
    "ford_usl_fbm", "ford_usl_fba", "ford_usl_web",
    "gm_usl_fbm", "gm_usl_fba", "gm_usl_web",
    "ford_fol_fbm", "ford_fol_fba", "ford_fol_web",
    "gm_fol_fbm", "gm_fol_fba", "gm_fol_web",
    "ram_fbm", "ram_fba", "ram_web",
]
ws.append(columns)

if SRC.exists():
    src = load_workbook(SRC, data_only=True)["Data and Forecast"]
    # Source columns: A=week, B-D=FordUSL, E-G=GMUSL, H-J=FordFOL, K-M=GMFOL, N-P=Ram, W=Type (col 23)
    for row in src.iter_rows(min_row=3, values_only=True):
        week = row[0]
        if week is None:
            continue
        rtype = row[22] if len(row) > 22 else None
        rtype = "Forecast" if rtype == "Forecast" else "Actual"
        vals = [row[i] if i < len(row) else None for i in range(1, 16)]
        vals = [v if isinstance(v, (int, float)) else None for v in vals]
        ws.append([week, rtype, *vals])

style_header(ws)
for cell in ws["A"][1:]:
    cell.number_format = "yyyy-mm-dd"
autosize(ws)

wb.save(OUT)
print(f"Wrote {OUT}")
